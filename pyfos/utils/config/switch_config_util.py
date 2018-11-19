#!/usr/bin/env python3

# Copyright 2018 Brocade Communications Systems LLC.  All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""

:mod:`switch_config_util` - PyFOS util for specific config op use case.
***********************************************************************************
The :mod:`switch_config_util` provides for specific config op use case.

This module is utility functions for switch config scripts.

"""

import json
import time
import os
import openpyxl
from pyfos import pyfos_util
from pyfos import pyfos_rest_util
import pyfos.pyfos_brocade_zone as pyfos_zone
from pyfos.utils.zoning import zoning_cfg_save


# pylint: disable=W0603
gr = 0


def get_envelope_name(ipaddr):
    return "FOS_" + ipaddr + "_" + time.strftime("%Y_%m_%d_%H_%M_%S")


def in_template(template, obj_name, attribute_field):
    if template is None:
        # generating a backup copy, not a template
        return True

    for obj in template:
        for k1, v1 in obj.items():
            if k1 == obj_name:
                if "all" in v1:
                    return True
                else:
                    return bool(attribute_field in v1)
    return False


def get_template(filename):
    fp = open(filename, 'r')
    template_txt = fp.read()
    template_dict = json.loads(template_txt)
    fp.close()
    return template_dict


def find_matching_attrib_entry(old_attrib, keys):
    for entry1 in old_attrib.value:
        found = 0
        for entry2 in old_attrib.value:
            for k_e, v_e in entry2.items():
                for key in keys:
                    for k, v in key.items():
                        if v_e.is_key:
                            if (k_e == k) and (v_e.getuservalue() == v):
                                found += 1
            if found == len(keys):
                return entry1
    return None


# pylint: disable=W0613
def process_patch_attribute_map(
        old_attrib, current_attrib, template, diff_set):
    patch_needed = 0
    old_v = old_attrib.getuservalue()
    current_v = current_attrib.getuservalue()

    if isinstance(current_v, list):
        for entry in current_attrib.value:
            keys = []
            for k, v in entry.items():
                if v.is_key:
                    temp_dict = {}
                    temp_dict[k] = v.getuservalue()
                    keys.append(temp_dict)
            matching_old_attrib_entry = find_matching_attrib_entry(
                    old_attrib, keys)
            if matching_old_attrib_entry:
                for k, v in entry.items():
                    if not v.is_key:
                        if k in matching_old_attrib_entry:
                            if (v.getuservalue() !=
                                    matching_old_attrib_entry[k].
                                    getuservalue()):
                                diff_set[CHANGED_D].append([
                                    False,
                                    current_attrib.name,
                                    keys,
                                    matching_old_attrib_entry[k].getuservalue(),
                                    v.getuservalue()])
                                patch_needed += 1
                                v.setuservalue(
                                    matching_old_attrib_entry[k].
                                    getuservalue())
    elif isinstance(current_v, dict):
        if not (old_attrib.is_empty() and current_attrib.is_empty()):
            if old_v != current_v:
                if current_attrib.is_config:
                    diff_set[CHANGED_D].append([
                        False,
                        current_attrib.name,
                        old_v,
                        current_v])
                    patch_needed += 1
                    current_attrib.setuservalue(old_v)
                else:
                    diff_set[CHANGED_D].append([
                        True,
                        current_attrib.name,
                        old_v,
                        current_v])
    else:
        print("unknown", current_v)

    return patch_needed


# pylint: disable=W0613
def process_post_attribute_map(
        old_attrib, current_attrib, template, diff_set):

    post_needed = 0
    old_v = old_attrib.getuservalue()
    current_v = current_attrib.getuservalue()

    if isinstance(old_v, list):
        index = 0
        for entry in old_attrib.value:
            keys = []
            for k, v in entry.items():
                if v.is_key:
                    temp_dict = {}
                    temp_dict[k] = v.getuservalue()
                    keys.append(temp_dict)

            matching_current_attrib_entry = find_matching_attrib_entry(
                    current_attrib, keys)
            if not matching_current_attrib_entry:
                diff_set[KEY_W_D_D].append(current_attrib.name)
                if current_attrib.is_config:
                    diff_set[DELETED_D].append([
                        False,
                        current_attrib.name,
                        keys,
                        old_v[index]])
                    post_needed += 1
                    current_attrib.value.append(entry)
                else:
                    diff_set[DELETED_D].append([
                        True,
                        current_attrib.name,
                        keys,
                        old_v[index]])
            else:
                current_attrib.value.remove(matching_current_attrib_entry)
            index += 1
    elif isinstance(current_v, dict):
        if not old_attrib.is_empty() and current_attrib.is_empty():
            diff_set[KEY_W_D_D].append(current_attrib.name)
            if current_attrib.is_config:
                diff_set[DELETED_D].append([
                    False,
                    current_attrib.name,
                    None,
                    old_v])
                post_needed += 1
                current_attrib.setuservalue(old_v)
            else:
                diff_set[DELETED_D].append([
                    True,
                    current_attrib.name,
                    None,
                    old_v])
    else:
        print("unknown")

    return post_needed


# pylint: disable=W0613
def process_delete_attribute_map(
        old_attrib, current_attrib, template,
        diff_set):
    delete_needed = 0
    old_v = old_attrib.getuservalue()
    current_v = current_attrib.getuservalue()

    ret_keys = []

    if current_attrib.is_empty():
        return delete_needed, ret_keys
        # by definition, if the current attribute is empty
        # there is nothing to be deleted to match to the
        # previously saved config

    if isinstance(current_v, list):
        index = 0
        for entry in current_attrib.value:
            keys = []
            for k, v in entry.items():
                if v.is_key:
                    temp_dict = {}
                    temp_dict[k] = v.getuservalue()
                    keys.append(temp_dict)

            matching_old_attrib_entry = find_matching_attrib_entry(
                    old_attrib, keys)
            if not matching_old_attrib_entry:
                diff_set[KEY_W_D_D].append(current_attrib.name)
                if old_attrib.is_config:
                    diff_set[ADDED_D].append([
                        False,
                        current_attrib.name,
                        keys,
                        current_v[index]])
                    delete_needed += 1
                    for key in keys:
                        ret_keys.append(key)
                else:
                    diff_set[ADDED_D].append([
                        True,
                        current_attrib.name,
                        keys,
                        current_v[index]])
            index += 1
    elif isinstance(current_v, dict):
        if not current_attrib.is_empty() and old_attrib.is_empty():
            diff_set[KEY_W_D_D].append(old_attrib.name)
            if old_attrib.is_config:
                diff_set[ADDED_D].append([
                    False,
                    old_attrib.name,
                    current_v])
                delete_needed += 1
                current_attrib.setuservalue(old_v)
            else:
                diff_set[ADDED_D].append([
                    True,
                    old_attrib.name,
                    current_v])
    else:
        print("unknown")

    return delete_needed, ret_keys


def process_patch(old_object, current_object, template, diff_set):
    keys = {}
    for k, v in current_object.attributes_dict.items():
        if v.is_key:
            keys[k] = v.getuservalue()

    print(
            "processing entries that changed",
            current_object.getcontainer(), keys)

    patch_needed = 0

    for k, v in current_object.attributes_dict.items():
        if v.is_key:
            # if the attribute is a key, no need to touch that
            continue
        elif v.is_attribute_map:
            local_patch_need = process_patch_attribute_map(
                    old_object.getattribute(k), v,
                    template, diff_set)
            patch_needed += local_patch_need
            continue
        elif v.is_list:
            print("deal with list later")
            continue
        elif v.is_leaf:
            if (template is not None and
                    not in_template(
                        template, current_object.getcontainer(), k)):
                continue

            if old_object.getattribute(k).getuservalue() != v.getuservalue():
                diff_set[KEY_W_D_D].append(k)
                if v.is_config:
                    diff_set[CHANGED_D].append([
                        False,
                        k,
                        old_object.getattribute(k).getuservalue(),
                        v.getuservalue()])
                    patch_needed += 1
                    v.setuservalue(
                        old_object.getattribute(k).getuservalue())
                else:
                    diff_set[CHANGED_D].append([
                        True,
                        k,
                        old_object.getattribute(k).getuservalue(),
                        v.getuservalue()])

    if patch_needed > 0:
        return True

    return False


def process_post(old_object, current_object, template, diff_set):
    keys = {}
    for k, v in current_object.attributes_dict.items():
        if v.is_key:
            keys[k] = v.getuservalue()

    print("processing entries requiring creation",
          current_object.getcontainer(), keys)

    post_needed = 0

    for k, v in current_object.attributes_dict.items():
        if v.is_key:
            # if the attribute is a key, no need to touch that
            continue
        elif v.is_attribute_map:
            local_post_need = process_post_attribute_map(
                    old_object.getattribute(k), v, template,
                    diff_set)
            post_needed += local_post_need
            continue
        elif v.is_list:
            print("deal with list later")
            continue
        elif v.is_leaf:
            # if both objects are there, this should be handled in patch
            continue

    if post_needed > 0:
        return True

    return False


def process_delete(
        old_object, current_object, delete_object, template,
        diff_set):
    keys = {}
    for k, v in current_object.attributes_dict.items():
        if v.is_key:
            keys[k] = v.getuservalue()

    print(
            "processing entries requiring deletion",
            current_object.getcontainer(), keys)

    delete_needed = 0

    for k, v in current_object.attributes_dict.items():
        if v.is_key:
            # if the attribute is a key, no need to touch that
            continue
        elif v.is_attribute_map:
            local_delete_need, keys = process_delete_attribute_map(
                    old_object.getattribute(k), v, template,
                    diff_set)
            if local_delete_need:
                setfunc = getattr(delete_object, "set_" + k.lower())
                setfunc(keys)
                delete_needed += local_delete_need
            continue
        elif v.is_list:
            print("deal with list later")
            continue
        elif v.is_leaf:
            # if both objects are there, this should be handled in patch
            continue

    if delete_needed > 0:
        return True

    return False


def find_matching_object(fos_object, old_object, refkeys):
    keys = []
    for k, v in fos_object.attributes_dict.items():
        if v.is_key:
            keys.append(k)

    if isinstance(old_object, list):
        # this should not be. If I have a list, there should be a key
        if len(keys) == 0:
            return None

        for obj in old_object:
            found = 0
            if refkeys is not None:
                for refkey in refkeys:
                    for k, v in refkey.items():
                        if obj.getattribute(k).getuservalue() == v:
                            found += 1
            else:
                for key in keys:
                    if (obj.getattribute(key).getuservalue() ==
                            fos_object.getattribute(key).getuservalue()):
                        found += 1
            if found == len(keys):
                return obj

        return None

    else:
        # if there are no keys associated here, just return
        # the single object
        if len(keys) == 0:
            return old_object

        found = 0
        if refkeys is not None:
            for refkey in refkeys:
                for k, v in refkey.items():
                    if obj.getattribute(k).getuservalue() == v:
                        found += 1
        else:
            for key in keys:
                if (old_object.getattribute(key).getuservalue() ==
                        fos_object.getattribute(key).getuservalue()):
                    found += 1

        if found == len(keys):
            return old_object
        else:
            return None


def obj_to_short_str(fos_object):
    values = []
    # pylint: disable=W0612
    for k, v in fos_object.attributes_dict.items():
        if v.is_key:
            values.append(v)

    return type(fos_object).__name__ + str(values)


def process_object(
        session, envelope_name, obj_handler,
        do_print, do_db_change, in_json, template, refkeys=None):

    pyfos_class = obj_handler["obj_name"]

    # get the current object
    fos_object = pyfos_class.get(session)
    if pyfos_util.is_failed_resp(fos_object):
        print("cannot retrieve from switch", pyfos_class.__name__)
        return

    # get the old object
    old_dict = None
    if in_json:
        file_name = envelope_name + "/" + pyfos_class.__name__
        try:
            os.stat(file_name)
        except OSError:
            print(file_name, "doesn't exist. Skipping")
            return

        fp = open(file_name, 'r')
        old_dict = json.load(fp)
        fp.close()
    else:
        file_name = envelope_name + ".xlsx"
        try:
            os.stat(file_name)
        except OSError:
            print(file_name, "doesn't exist. Skipping")
            return

        reader = None
        if "reader" in obj_handler:
            reader = obj_handler["reader"]
        else:
            reader = read_simple_object
        old_dict = reader(file_name, pyfos_class)

    if old_dict is None:
        print("cannot read from file", pyfos_class.__name__)
        return

    compare_old_dict_and_new_object(
        session, pyfos_class, old_dict, fos_object, refkeys, template,
        do_print, do_db_change, obj_handler)


KEY_W_D_D = "keys_with_diff_data"
ADDED_D = "added_data"
DELETED_D = "deleted_data"
CHANGED_D = "changed_data"


def diff_set_print(diff_set, print_type):
    if print_type is None or print_type is ADDED_D:
        for diff in diff_set[ADDED_D]:
            if diff[0] is False:
                print("\t", end="")
            else:
                print("\t(RO)", end=" ")
            if len(diff) == 4:
                print(diff[1], diff[2], "Added. New value", diff[3])
            elif len(diff) == 3:
                print(diff[1], "Added. New value", diff[2])
            else:
                print("invalid length", len(diff))

    if print_type is None or print_type is DELETED_D:
        for diff in diff_set[DELETED_D]:
            if diff[0] is False:
                print("\t", end="")
            else:
                print("\t(RO)", end=" ")
            if len(diff) == 4:
                print(diff[1], diff[2], "removed. Old value", diff[3])
            elif len(diff) == 3:
                print(diff[1], "removed. Old value", diff[2])
            else:
                print("invalid length", len(diff))

    if print_type is None or print_type is CHANGED_D:
        for diff in diff_set[CHANGED_D]:
            if diff[0] is False:
                print("\t", end="")
            else:
                print("\t(RO)", end=" ")
            if len(diff) == 5:
                print(diff[1], diff[2],
                      "drifted from", diff[3],
                      "to", diff[4])
            elif len(diff) == 4:
                print(diff[1], "drifted from",
                      diff[2], "to", diff[3])
            else:
                print("invalid length", len(diff))


def compare_old_dict_and_new_object(session, pyfos_class, old_dict,
                                    fos_object, refkeys, template, do_print, do_db_change, obj_handler):

    changed = False

    diff_set = {}
    diff_set[KEY_W_D_D] = []
    diff_set[ADDED_D] = []
    diff_set[DELETED_D] = []
    diff_set[CHANGED_D] = []

    old_object = None
    if isinstance(old_dict, list):
        old_object = []
        for old_entry in old_dict:
            old_entry_obj = pyfos_class()
            old_entry_obj.load(old_entry[pyfos_class().getcontainer()])
            old_object.append(old_entry_obj)
    else:
        old_object = pyfos_class()
        old_object.load(old_dict[pyfos_class().getcontainer()])

        # if isinstance(old_object, list):
        #   for obj in old_object:
        #       print(json.dumps(obj,
        #           cls=pyfos_rest_util.rest_object_encoder,
        #           sort_keys=True, indent=4))
        # else:
        #   print(json.dumps(old_object,
        #           cls=pyfos_rest_util.rest_object_encoder,
        #           sort_keys=True, indent=4))

    if isinstance(fos_object, list):
        for obj in fos_object:
            old_matching_object = find_matching_object(
                    obj, old_object, refkeys)
            if old_matching_object is None:
                print("didn't find a matching old object",
                      obj_to_short_str(obj))
            else:
                # print("compare")
                # print(json.dumps(obj,
                # cls=pyfos_rest_util.rest_object_encoder,
                # sort_keys=True, indent=4))
                # print(json.dumps(old_matching_object,
                # cls=pyfos_rest_util.rest_object_encoder,
                # sort_keys=True, indent=4))
                need_patching = process_patch(
                        old_matching_object, obj, template,
                        diff_set)
                if do_print:
                    diff_set_print(diff_set, CHANGED_D)
                if do_db_change and need_patching:
                    if "during_patch" in obj_handler and obj_handler["during_patch"] is not None:
                        obj_handler["during_patch"](session, obj)
                    result = obj.patch(session)
                    if ('success-type' in result and
                            result['success-type'] == 'Success'):
                        changed = True
                        print("\tpatch success")
                    else:
                        print("\tpatch failed", result)
    else:
        old_matching_object = find_matching_object(
                fos_object, old_object, refkeys)
        if old_matching_object is None:
            print("didn't find a matching old object",
                  obj_to_short_str(fos_object))
        else:

            # print("compare")
            # print(json.dumps(fos_object,
            #       cls=pyfos_rest_util.rest_object_encoder,
            #       sort_keys=True, indent=4))
            # print(json.dumps(old_matching_object,
            #       cls=pyfos_rest_util.rest_object_encoder,
            #       sort_keys=True, indent=4))
            need_patching = process_patch(
                    old_matching_object, fos_object, template,
                    diff_set)
            if do_print:
                diff_set_print(diff_set, CHANGED_D)
            if do_db_change and need_patching:
                if "during_patch" in obj_handler and obj_handler["during_patch"] is not None:
                    obj_handler["during_patch"](session, fos_object)
                # print(fos_object.create_html_content(pyfos_rest_util.rest_get_method.GET_KEY_AND_MODIFIED_CONFIG))
                result = fos_object.patch(session)
                if ('success-type' in result and
                        result['success-type'] == 'Success'):
                    changed = True
                    print("\tpatch success")
                else:
                    print("\tpatch failed", result)

            need_posting = process_post(
                    old_matching_object, fos_object, template,
                    diff_set)
            if do_print:
                diff_set_print(diff_set, DELETED_D)

            if do_db_change and need_posting:
                # print(fos_object.create_html_content())
                result = fos_object.post(session)
                if ('success-type' in result and
                        result['success-type'] == 'Success'):
                    changed = True
                    print("\tpost success")
                else:
                    print("\tpost failed", result)

            delete_object = pyfos_class()
            need_deleting = process_delete(
                    old_matching_object, fos_object,
                    delete_object, template, diff_set)
            if do_print:
                diff_set_print(diff_set, ADDED_D)

            if do_db_change and need_deleting:
                # print(delete_object.create_html_content(pyfos_rest_util.rest_get_method.GET_KEY_AND_MODIFIED_CONFIG_DELETE))
                result = delete_object.delete(session)
                if ('success-type' in result and
                        result['success-type'] == 'Success'):
                    changed = True
                    print("\tdelete success")
                else:
                    print("\tdelete failed", result)

    if do_db_change:
        if "after_patch_post_delete" in obj_handler:
            post_changed_handler = obj_handler["after_patch_post_delete"]
            if changed and post_changed_handler is not None:
                post_changed_handler(session, diff_set)

    return diff_set


def process_list(ws, fos_list, sc):
    global gr

    for obj1 in fos_list:
        if isinstance(obj1, list):
            process_list(ws, obj1, sc)
        elif isinstance(obj1, dict):
            process_dict(ws, obj1, sc)
        else:
            ws.cell(column=sc, row=gr, value=obj1)
            gr = gr + 1


def process_dict(ws, fos_dict, sc):
    global gr

    for k, v in fos_dict.items():
        if isinstance(v, list):
            for obj1 in v:
                if isinstance(obj1, list):
                    process_list(ws, obj1, sc)
                elif isinstance(obj1, dict):
                    process_dict(ws, obj1, sc)
                else:
                    ws.cell(column=sc, row=gr, value=obj1)
                    gr = gr + 1
        elif isinstance(v, dict):
            ws.cell(column=sc, row=gr, value=k)
            gr = gr + 1
            process_dict(ws, v, sc + 1)
        else:
            ws.cell(column=sc, row=gr, value=k)
            ws.cell(column=sc + 1, row=gr, value=v)
            gr = gr + 1


def dump_object_in_json(session, pyfos_class, dir_name):
    fos_object = pyfos_class.get(session)

    fp = open(dir_name + "/" + pyfos_class.__name__, 'w')

    fp.write(json.dumps(
        fos_object,
        cls=pyfos_rest_util.rest_object_encoder,
        sort_keys=True, indent=4))

    fp.close()

    print("dumped", pyfos_class.__name__)


def process_simple_dict(ws, parent, fos_dict, title_dict, sr):
    rows_taken = 0
    rows_by_children = 0
    for k, v in fos_dict.items():
        if isinstance(v, list):
            child_rows = 0
            for obj1 in v:
                if isinstance(obj1, list):
                    print("invalid definition. list within list", k, v)
#                    process_simple_list(ws, k, obj1, title_dict, sr)
                elif isinstance(obj1, dict):
                    child_rows = process_simple_dict(ws, k, obj1, title_dict, sr)
                else:
                    if parent is not None:
                        sc = title_dict[parent + "." + k]
                    else:
                        sc = title_dict[k]
                    ws.cell(column=sc, row=sr, value=obj1)
                    sr = sr + 1
                    child_rows = child_rows + 1

            if child_rows > 0 and rows_by_children < child_rows:
                rows_by_children = child_rows

        elif isinstance(v, dict):
            child_rows = process_simple_dict(ws, k, v, title_dict, sr)

            if child_rows > 0 and rows_by_children < child_rows:
                rows_by_children = child_rows
        else:
            if parent is not None:
                sc = title_dict[parent + "." + k]
            else:
                sc = title_dict[k]
            ws.cell(column=sc, row=sr, value=v)

    if rows_by_children > 0:
        rows_taken = rows_taken + rows_by_children
    else:
        rows_taken = 1

    # let's leave a space between two rows
    rows_taken = rows_taken + 1

    return rows_taken


def create_header_row(ws, pyfos_class, title_dict):
    fos_object = pyfos_class()

    sc = 1
    for attribute in fos_object.attributes:
        ws.cell(column=sc, row=1, value=attribute.name)
        title_dict[attribute.name] = sc
        sc = sc + 1
        if isinstance(attribute.getvalue(), list):
            print("attribute value list", attribute.getvalue())
            for elem in attribute.getvalue():
                if isinstance(elem, dict):
                    for k, v in elem.items():
                        if isinstance(v, dict):
                            # pylint: disable=W0612
                            for k1, v1 in elem.items():
                                heading_name = attribute.name + "." + k + "." + k1
                                ws.cell(column=sc, row=1, value=heading_name)
                                title_dict[heading_name] = sc
                                print("adding ", heading_name)
                                sc = sc + 1
                        else:
                            heading_name = attribute.name + "." + k
                            ws.cell(column=sc, row=1, value=heading_name)
                            title_dict[heading_name] = sc
                            print("adding ", heading_name)
                            sc = sc + 1
                else:
                    print("don't know how to handle", elem)
        elif isinstance(attribute.getvalue(), dict):
            for k, v in attribute.getvalue().items():
                heading_name = attribute.name + "." + k
                ws.cell(column=sc, row=1, value=heading_name)
                title_dict[heading_name] = sc
                sc = sc + 1
# title the rest
#    for attribute in fos_object.attributes:
#        if attribute.version_supported.visible(fos_object.version_active):
#            if attribute.is_key is not rest_yang_config.yang_key:
#                ws.cell(column=sc, row=1, value=attribute.name)
#                title_dict[attribute.name] = sc
#                sc = sc + 1


def write_simple_object(session, pyfos_class, wb):
    global gr

    fos_object = pyfos_class.get(session)
    if pyfos_util.is_failed_resp(fos_object):
        print("failed to dump", pyfos_class.__name__)
        return

    ws = wb.create_sheet(pyfos_class.__name__, 0)

    fos_data = json.loads(json.dumps(
        fos_object,
        cls=pyfos_rest_util.rest_object_encoder,
        sort_keys=True, indent=4))

    sr = 2
    title_dict = {}

    create_header_row(ws, pyfos_class, title_dict)

    if isinstance(fos_data, list):

        for data in fos_data:
            if isinstance(data, dict):
                # pylint: disable=W0612
                for k, v in data.items():
                    rows_taken = process_simple_dict(ws, None, v, title_dict, sr)
                    sr = sr + rows_taken
            else:
                print("shouldn't be non-dict", data)
    elif isinstance(fos_data, dict):

        for k, v in fos_data.items():
            rows_taken = process_simple_dict(ws, None, v, title_dict, sr)
            sr = sr + rows_taken
    else:
        print("not sure why", type(fos_object))

    print("dumped", pyfos_class.__name__)


def write_defined_zone_object(session, pyfos_class, wb):
    global gr

    fos_object = pyfos_class.get(session)
    if pyfos_util.is_failed_resp(fos_object):
        print("failed to dump", pyfos_class.__name__)
        return

    ws = wb.create_sheet(pyfos_class.__name__, 0)

    json.loads(json.dumps(
        fos_object,
        cls=pyfos_rest_util.rest_object_encoder,
        sort_keys=True, indent=4))

    ws.cell(column=1, row=1, value="cfg")
    ws.cell(column=1, row=2, value="cfg-name")
    ws.cell(column=2, row=2, value="zone-name")
    cfgs = fos_object.peek_cfg()
    sr = 3
    for cfg in cfgs:
        ws.cell(column=1, row=sr, value=cfg['cfg-name'])
        sc = 2
        for zone in cfg['member-zone']['zone-name']:
            ws.cell(column=sc, row=sr, value=zone)
            sc = sc + 1
        sr = sr + 1

    aliases = fos_object.peek_alias()
    # skip a line
    sr = sr + 1
    ws.cell(column=1, row=sr, value="alias")
    sr = sr + 1
    ws.cell(column=1, row=sr, value="alias-name")
    ws.cell(column=2, row=sr, value="alias-entry-name")
    sr = sr + 1
    for alias in aliases:
        ws.cell(column=1, row=sr, value=alias['alias-name'])
        sc = 2
        for member in alias['member-entry']['alias-entry-name']:
            ws.cell(column=sc, row=sr, value=member)
            sc = sc + 1
        sr = sr + 1

    zones = fos_object.peek_zone()

    # skip a line
    sr = sr + 1
    ws.cell(column=1, row=sr, value="zone")
    sr = sr + 1
    ws.cell(column=1, row=sr, value="zone-name")
    ws.cell(column=2, row=sr, value="zone-type")
    ws.cell(column=3, row=sr, value="entry-name")
    sr = sr + 1
    ws.cell(column=3, row=sr, value="principal-entry-name")
    sr = sr + 1
    for zone in zones:
        ws.cell(column=1, row=sr, value=zone['zone-name'])
        ws.cell(column=2, row=sr, value=zone['zone-type'])
        sc = 3
        psc = 3
        for member in zone['member-entry']['entry-name']:
            ws.cell(column=sc, row=sr, value=member)
            sc = sc + 1
        sr = sr + 1
        for member in zone['member-entry']['principal-entry-name']:
            ws.cell(column=psc, row=sr, value=member)
            psc = psc + 1
        sr = sr + 1

    print("dumped", pyfos_class.__name__)


def write_effective_zone_object(session, pyfos_class, wb):
    global gr

    fos_object = pyfos_class.get(session)
    if pyfos_util.is_failed_resp(fos_object):
        print("failed to dump", pyfos_class.__name__)
        return

    ws = wb.create_sheet(pyfos_class.__name__, 0)

    json.loads(json.dumps(
        fos_object,
        cls=pyfos_rest_util.rest_object_encoder,
        sort_keys=True, indent=4))

    ws.cell(column=1, row=1, value="cfg-name")
    ws.cell(column=1, row=2, value=fos_object.peek_cfg_name())
    ws.cell(column=2, row=1, value="checksum")
    ws.cell(column=2, row=2, value=fos_object.peek_checksum())
    ws.cell(column=3, row=1, value="cfg-action")
    ws.cell(column=3, row=2, value=fos_object.peek_cfg_action())
    ws.cell(column=4, row=1, value="default-zone-access")
    ws.cell(column=4, row=2, value=fos_object.peek_default_zone_access())
    ws.cell(column=5, row=1, value="db-avail")
    ws.cell(column=5, row=2, value=fos_object.peek_db_avail())
    ws.cell(column=6, row=1, value="db-max")
    ws.cell(column=6, row=2, value=fos_object.peek_db_max())
    ws.cell(column=7, row=1, value="db-committed")
    ws.cell(column=7, row=2, value=fos_object.peek_db_committed())
    ws.cell(column=8, row=1, value="db-transaction")
    ws.cell(column=8, row=2, value=fos_object.peek_db_transaction())
    ws.cell(column=9, row=1, value="transaction-token")
    ws.cell(column=9, row=2, value=fos_object.peek_transaction_token())
    ws.cell(column=9, row=1, value="db-chassis-wide-committed")
    ws.cell(column=9, row=2, value=fos_object.peek_db_chassis_wide_committed())

    ws.cell(column=1, row=4, value="enabled-zone")
    sr = 5
    zones = fos_object.peek_enabled_zone()

    # skip a line
    sr = sr + 1
    ws.cell(column=1, row=sr, value="zone")
    sr = sr + 1
    ws.cell(column=1, row=sr, value="zone-name")
    ws.cell(column=2, row=sr, value="zone-type")
    ws.cell(column=3, row=sr, value="entry-name")
    sr = sr + 1
    ws.cell(column=3, row=sr, value="principal-entry-name")
    sr = sr + 1
    for zone in zones:
        ws.cell(column=1, row=sr, value=zone['zone-name'])
        ws.cell(column=2, row=sr, value=zone['zone-type'])
        sc = 3
        psc = 3
        for member in zone['member-entry']['entry-name']:
            ws.cell(column=sc, row=sr, value=member)
            sc = sc + 1
        sr = sr + 1
        for member in zone['member-entry']['principal-entry-name']:
            ws.cell(column=psc, row=sr, value=member)
            psc = psc + 1
        sr = sr + 1

    print("dumped", pyfos_class.__name__)


def read_simple_object(file_name, pyfos_class):
    wb = openpyxl.load_workbook(file_name)
    sheet_name = pyfos_class.__name__

    if sheet_name not in wb.sheetnames:
        print("worksheet is not present", sheet_name)
        return None

    ws = wb[sheet_name]

    num_of_entries = 0
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None:
            num_of_entries = num_of_entries + 1

    if num_of_entries is 1:
        old_dict = {}
    elif num_of_entries > 1:
        old_dict = []
    else:
        print("wrong format sheet", pyfos_class.__name__)
        return None

    fos_object = pyfos_class()

    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue

        each_row = {}
        each_row[fos_object.getcontainer()] = {}
        for cell in row:
            header = ws.cell(column=cell.col_idx, row=1).value
            if "." in header:
                container, k = header.split(".")
                if container in each_row[fos_object.getcontainer()]:
                    if not isinstance(each_row[fos_object.getcontainer()][container], dict):
                        each_row[fos_object.getcontainer()][container] = {}
                else:
                    print("container should be here", container)

                mylist = []
                for rindex in range(cell.row, ws.max_row + 1):
                    list_cell = ws.cell(column=cell.col_idx, row=rindex)
                    if list_cell.value is None:
                        break
                    mylist.append(list_cell.value)
                if len(mylist) > 0:
                    if len(mylist) == 1:
                        each_row[fos_object.getcontainer()][container][k] = mylist[0]
                    else:
                        each_row[fos_object.getcontainer()][container][k] = mylist
            else:
                each_row[fos_object.getcontainer()][header] = cell.value
        if isinstance(old_dict, list):
            old_dict.append(each_row)
        else:
            old_dict = each_row

    return old_dict


ZONE1 = 1
ZONE2 = 2
CFG = 3
ALIAS = 4


def read_defined_zone_object(file_name, pyfos_class):
    wb = openpyxl.load_workbook(file_name)
    sheet_name = pyfos_class.__name__

    ws = wb[sheet_name]

    fos_object = pyfos_class()

    old_dict = {}
    old_dict[fos_object.getcontainer()] = {}

    zones = []

    cfgs = []

    aliases = []

    process_next = None

    zone = {}
    alias = {}
    cfg = {}
    for row in ws.iter_rows():
        if row[0].value == "zone-name":
            process_next = ZONE1
        elif row[0].value == "cfg-name":
            process_next = CFG
        elif row[0].value == "alias-name":
            process_next = ALIAS
        elif row[0].value == "zone" or row[0].value == "cfg" or row[0].value == "alias":
            continue
        elif row[0].value is None:
            if process_next is ZONE2:
                pmembers = []
                for cell in row:
                    if (cell.col_idx > 2) and cell.value is not None:
                        pmembers.append(cell.value)
                zone["member-entry"]["principal-entry-name"] = pmembers
                zones.append(zone)
                process_next = ZONE1
            else:
                continue
        else:
            if process_next is ALIAS:
                alias = {}
                alias["alias-name"] = row[0].value
                members = []
                for cell in row:
                    if (cell.col_idx > 1) and cell.value is not None:
                        members.append(cell.value)
                alias["member-entry"] = {}
                alias["member-entry"]["alias-entry-name"] = members
                aliases.append(alias)
            elif process_next is CFG:
                cfg = {}
                cfg["cfg-name"] = row[0].value
                members = []
                for cell in row:
                    if (cell.col_idx > 1) and cell.value is not None:
                        members.append(cell.value)
                cfg["member-zone"] = {}
                cfg["member-zone"]["zone-name"] = members
                cfgs.append(cfg)
            elif process_next is ZONE1:
                zone = {}
                zone["zone-name"] = row[0].value
                zone["zone-type"] = row[1].value
                members = []
                for cell in row:
                    if (cell.col_idx > 2) and cell.value is not None:
                        members.append(cell.value)
                zone["member-entry"] = {}
                zone["member-entry"]["entry-name"] = members
                process_next = ZONE2
            else:
                print("shouldn't be here")

    # empty principal zone members in the last gets skipped
    if process_next is ZONE2:
        zones.append(zone)

    old_dict[fos_object.getcontainer()]["alias"] = []
    for alias in aliases:
        old_dict[fos_object.getcontainer()]["alias"].append(alias)
    old_dict[fos_object.getcontainer()]["cfg"] = []
    for cfg in cfgs:
        old_dict[fos_object.getcontainer()]["cfg"].append(cfg)
    old_dict[fos_object.getcontainer()]["zone"] = []
    for zone in zones:
        old_dict[fos_object.getcontainer()]["zone"].append(zone)

    return old_dict


def read_effective_zone_object(file_name, pyfos_class):
    wb = openpyxl.load_workbook(file_name)
    sheet_name = pyfos_class.__name__

    ws = wb[sheet_name]

    fos_object = pyfos_class()

    old_dict = {}
    old_dict[fos_object.getcontainer()] = {}

    old_dict[fos_object.getcontainer()]["cfg-name"] = ws.cell(column=1, row=2).value
    old_dict[fos_object.getcontainer()]["checksum"] = ws.cell(column=2, row=2).value
    old_dict[fos_object.getcontainer()]["cfg-action"] = ws.cell(column=3, row=2).value
    old_dict[fos_object.getcontainer()]["default-zone-access"] = ws.cell(column=4, row=2).value
    old_dict[fos_object.getcontainer()]["db-avail"] = ws.cell(column=5, row=2).value
    old_dict[fos_object.getcontainer()]["db-max"] = ws.cell(column=6, row=2).value
    old_dict[fos_object.getcontainer()]["db-committed"] = ws.cell(column=7, row=2).value
    old_dict[fos_object.getcontainer()]["db-transaction"] = ws.cell(column=8, row=2).value
    old_dict[fos_object.getcontainer()]["transaction-token"] = ws.cell(column=9, row=2).value
    old_dict[fos_object.getcontainer()]["db-chassis-wide-committed"] = ws.cell(column=9, row=2).value

    zones = []

    process_next = ZONE1

    zone = {}
    for row in ws.iter_rows(min_row=9):
        if process_next is ZONE2:
            pmembers = []
            for cell in row:
                if (cell.col_idx > 2) and cell.value is not None:
                    pmembers.append(cell.value)
            zone["member-entry"]["principal-entry-name"] = pmembers
            zones.append(zone)
            process_next = ZONE1
        elif process_next is ZONE1:
            zone = {}
            zone["zone-name"] = row[0].value
            zone["zone-type"] = row[1].value
            members = []
            for cell in row:
                if (cell.col_idx > 2) and cell.value is not None:
                    members.append(cell.value)
            zone["member-entry"] = {}
            zone["member-entry"]["entry-name"] = members
            process_next = ZONE2
        else:
            print("shouldn't be here")

    # empty principal zone members in the last gets skipped
    if process_next is ZONE2:
        zones.append(zone)

    old_dict[fos_object.getcontainer()]["enabled-zone"] = []
    for zone in zones:
        old_dict[fos_object.getcontainer()]["enabled-zone"].append(zone)

    return old_dict


# pylint: disable=W0613
def save_change_defined_zone_object(session, diff_set):
    current_effective = pyfos_zone.effective_configuration.get(session)
    result = zoning_cfg_save.cfgsave(session, current_effective.peek_checksum())
    if ('success-type' in result and result['success-type'] == 'Success'):
        print("\tsave_change_defined_zone_object success")
    else:
        print("\tsave_change_defined_zone_object failed", result)


def during_patch_effective_zone_object(session, fos_obj):
    cfg_name_attribute = fos_obj.getattribute("cfg-name")
    if cfg_name_attribute is not None and cfg_name_attribute.getisvaluechanged():
        current_effective = pyfos_zone.effective_configuration.get(session)
        fos_obj.set_checksum(current_effective.peek_checksum())
        if fos_obj.peek_cfg_name() is None:
            fos_obj.set_cfg_action(pyfos_zone.CFG_ACTION_DISABLE)
    else:
        checksum_attribute = fos_obj.getattribute("checksum")
        if checksum_attribute is not None:
            checksum_attribute.setvaluechanged(0)


def save_change_effective_zone_object(session, diff_set):
    if "default-zone-access" in diff_set[KEY_W_D_D]:
        current_effective = pyfos_zone.effective_configuration.get(session)
        result = zoning_cfg_save.cfgsave(session, current_effective.peek_checksum())
        if ('success-type' in result and result['success-type'] == 'Success'):
            print("\tsave_change_effective_zone_object success")
        else:
            print("\tsave_change_effective_zone_object failed", result)
