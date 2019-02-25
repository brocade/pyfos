# Copyright 2018 Brocade Communications Systems LLC.  All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.


"""

:mod:`pyfos_config_manager` - PyFOS module  to provide config related\
        handling Support.
*************************************************************************************************
The :mod:`pyfos_config_manager`  Provides the Config functionality for REST
supported classes.

"""


import json
import re
import os
from pathlib import Path
import openpyxl
import pyfos.pyfos_util as utils
from pyfos.manager.pyfos_class_manager import clsmanager
from pyfos.manager.pyfos_rule_manager import rmmanager
from pyfos.pyfos_brocade_fibrechannel_logical_switch import \
    fibrechannel_logical_switch
from pyfos.pyfos_brocade_fru import blade


class config_manager():
    """
    class for config management for the rest supported classes in the infra.
    """
    def __init__(self, fmtfile="XLSX", fmtobj="attributes"):
        self.version = "1.1"
        if fmtfile in ['XLSX', 'JSON']:
            self.fmtfile = fmtfile
        else:
            self.fmtfile = 'XLSX'

        if fmtobj in ['json', 'attributes']:
            self.fmtobj = fmtobj
        else:
            self.fmtobj = 'attributes'

        # init it
        clsmanager.clsglobalinit(1)
        self.clslist = clsmanager.getAllCls()
        self.rmmanager = None
        self.current = None
        self.loaded = None
        # dump config
        self.dumplist = list()
        self.dumpfids = list()
        self.dumpslots = list()
        self.dumpcontainers = list()
        # load config
        self.loadfids = list()
        self.loaddict = dict()
        self.loadslots = list()
        # diff config
        self.difffids = list()
        self.diffdict = dict()
        self.diffdetails = dict()
        self.diffslots = list()
        # current config
        self.currentfids = list()
        self.currentdict = dict()
        self.currentslots = list()
        self.chunkcount = 1
        self.columncount = 20
        # filters
        self.clsfilters = []
        self.logger = clsmanager.initlogger('ConfigManager')
        self.configlist = list()
        self.loadconfiglist()

    def loadconfiglist(self):
        """
        Load the config supported class list
        """
        configfileanme = Path("configfilters.json")
        if configfileanme.exists():
            Confighandle = open(configfileanme.name, 'r')
            configdata = Confighandle.read()
            Confighandle.close()
            try:
                self.configlist = json.loads(configdata)
            except ValueError as err:
                print("Unable to load filters : ", err)
                self.log(3, "Unable to load filters : ", err)
        # print (self.loaddict)
        return 0

    def log(self, level, *msg):
        clsmanager.logbase(self.logger, level, "".join(map(str, msg)))

    def dumptofile(self, session, dumpfile, fmtfile=None, fmtobj=None):
        """Dump the switch configuration to a persistent file."""
        print("Dump configuration Start")
        self.log(1, "Dump configuration Start")
        if fmtfile is None:
            fmtfile = self.fmtfile
        if fmtobj is None:
            fmtobj = self.fmtobj

        ret = blade.get(session)
        ret = fibrechannel_logical_switch.get(session)
        if utils.is_failed_resp(ret):
            # confirm ls is not supported for platform like AG and Awing
            fid = "None"
            self.dumpformatfid(session, fid, dumpfile, fmtfile, fmtobj)
            if fid not in self.dumpfids:
                self.dumpfids.append(fid)
        else:
            if not isinstance(ret, list):
                ret = [ret]
            lslist = ret
            for i in range(len(lslist)):
                fid = str(lslist[i].peek_fabric_id())
                self.dumpformatfid(session, fid, dumpfile, fmtfile, fmtobj)
                if fid not in self.dumpfids:
                    self.dumpfids.append(fid)
        print("Dump configuration Completed")
        self.log(1, "Dump configuration Completed")

    def dumpheader(self, wb, fmtfile, fmtobj):
        """Dump the Header configuration to a persistent file."""
        ws = wb.create_sheet("DUMPHEADER")
        ws['A1'] = json.dumps(self.version)
        ws['A2'] = json.dumps(fmtfile)
        ws['A3'] = json.dumps(fmtobj)

    def dumpformatfid(self, session, fid, dumpfile, fmtfile, fmtobj):
        """Dump the switch configuration objects to a persistent file."""
        oldfid = session['vfid']
        if fmtfile in 'XLSX':
            wb = openpyxl.Workbook()
            if fid != "None":
                session['vfid'] = fid
            self.dumptofidclasses(wb, session, fmtobj)
            fidfile = str(dumpfile)
            fidfile = re.sub(".xlsx", "_" + str(fid) + ".xlsx", fidfile)
            self.dumpheader(wb, fmtfile, fmtobj)
            del wb['Sheet']
            wb.save(filename=fidfile)
            print("Saving Dump file", fidfile)
            self.log(1, "Saving Dump file", fidfile)
        else:
            fidfile = str(dumpfile)
            fidfile = re.sub(".json", "_" + str(fid) + ".json", fidfile)
            wb = open(fidfile, 'w')

            if fid != "None":
                session['vfid'] = fid
            self.dumptofiddefault(wb, session, fmtobj)
            wb.close()
            print("Saving Dump file", fidfile)
            self.log(1, "Saving Dump file", fidfile)
        session['vfid'] = oldfid

    def dumptofiddefault(self, wb, session, fmtobj):
        """
        Dump the switch configuration object in .xlsx file in JSON format.
        """
        self.dumplist = []
        if fmtobj != 'json':
            self.log(1, "Defaulting to json for the supported fmtfile")
            print("Defaulting to json for the supported fmtfile")
        for i in range(len(self.clslist)):
            cls = self.clslist[i]
            if any(self.configlist) and\
               cls().getcontainer() not in self.configlist:
                continue
            if not clsmanager.isConfigClass(cls, session):
                # print("skipping non config class", self.clslist[i])
                continue
            ret = self.clslist[i].get(session)
            if utils.is_failed_resp(ret):
                # print(ret)
                continue
            if not isinstance(ret, list):
                ret = [ret]
            self.dumplist += ret
        wb.write(str(self.dumplist))

    def dumptofidclasses(self, wb, session, fmtobj):
        """
        Dump the switch configuration object in .xlsx file based on format.
        """
        containerlist = []
        for i in range(len(self.clslist)):
            cls = self.clslist[i]
            if not clsmanager.isConfigClass(cls, session):
                # print("skipping non config class", self.clslist[i])
                continue
            ret = self.clslist[i].get(session)
            if utils.is_failed_resp(ret):
                # print(ret)
                continue
            if not isinstance(ret, list):
                ret = [ret]
            container = clsmanager.getContainerfromCls(cls)
            containerlist += [container]
            sheetdict = clsmanager.getmultiplesheetdict(container)
            if sheetdict is None or fmtobj == 'json':
                ws = wb.create_sheet(container)
            self.dumplist = ret
            listlen = len(self.dumplist)
            if fmtobj == 'json':
                chunk = self.chunkcount
                cols = self.columncount
                loopcount = int(listlen / chunk) + 1
                if chunk > 1 and int(listlen % chunk) > 0:
                    loopcount += 1
                endindex = 0
                for j in range(loopcount):
                    startindex = endindex
                    endindex = startindex + chunk
                    r = int(2 + (j / cols))
                    c = int(2 + (j % cols))
                    cell = ws.cell(row=r, column=c)
                    if loopcount >= j+1:
                        cell.value = str(self.dumplist[startindex:endindex])
                    else:
                        cell.value = str(self.dumplist[startindex:])
            else:
                baserow = 5
                addcount = 0
                for j in range(len(self.dumplist)):
                    obj = self.dumplist[j]
                    if sheetdict is None:
                        if j == 0:
                            obj.dumpheaders(ws, baserow - 1)
                        count = obj.dumpdata(ws, baserow + addcount)
                        addcount += count
                    else:
                        for k, v in sheetdict.items():
                            # print(k, v)
                            ws = wb.create_sheet(container + "." + k)
                            if j == 0:
                                obj.dumpheaders(ws, baserow - 1, v)
                            count = obj.dumpdata(ws,
                                                 baserow +
                                                 addcount,
                                                 v)
            # cell = ws.cell(row=1, column=1)
            # cell.value = listlen
        self.dumpcontainers = containerlist

    def loadcurrent(self, session, container=None):
        """Load the current switch configuration."""
        ret = blade.get(session)
        self.currentslots = clsmanager.getbladeslots(ret)
        ret = fibrechannel_logical_switch.get(session)
        if utils.is_failed_resp(ret):
            # confirm ls is not supported for platform like AG and Awing
            fid = "None"
            if fid not in self.currentfids:
                self.currentfids.append(fid)
            self.loadcurrentfid(session, fid, container)
        else:
            if not isinstance(ret, list):
                ret = [ret]
            lslist = ret
            oldfid = session['vfid']
            for i in range(len(lslist)):
                fid = str(lslist[i].peek_fabric_id())
                # print (fid, type(fid))
                if fid != "None":
                    session['vfid'] = fid
                self.loadcurrentfid(session, fid, container)
                if fid not in self.currentfids:
                    self.currentfids.append(fid)
            session['vfid'] = oldfid
        # print(self.currentdict)
        self.log(1, self.currentdict)

    def loadcurrentfid(self, session, fid, container):
        """Load the current switch configuration per fid."""
        print("Loading Switch configuration Start[", fid, "].")
        self.log(1, "Loading Switch configuration Start[", fid, "].")
        tmpdict = dict()
        for i in range(len(self.clslist)):
            if not clsmanager.isConfigClass(self.clslist[i], session):
                continue
            if container is not None and \
               container != self.clslist[i]().getcontainer():
                continue
            if any(self.configlist) and\
               self.clslist[i]().getcontainer() not in self.configlist:
                continue
            ret = self.clslist[i].get(session)
            if utils.is_failed_resp(ret):
                # print(ret)
                tmpdict.update(dict({self.clslist[i]().getcontainer():
                                     list()}))
                continue
            if not isinstance(ret, list):
                ret = [ret]
            tmpdict.update(dict({self.clslist[i]().getcontainer(): ret}))
        self.currentdict.update(dict({fid: tmpdict}))
        print("Loading Switch configuration Complete.")
        self.log(1, "Loading Switch configuration Complete.")

    def loadfromfile(self, dumpfile, fmtfile=None, fmtobj=None, gob=None):
        """Load the switch configuration from a dump file."""
        print("Loading Dump configuration Start.")
        self.log(1, "Loading Dump configuration Start.")
        # pylint: disable=W1401
        ext = re.sub(".*\.", ".", str(dumpfile))
        if ext == '.json':
            if fmtfile is None:
                fmtfile = 'JSON'
            elif fmtfile == 'XLSX':
                self.log(3, "Incorrect format specified for object load")
                print("Incorrect format specified for object load")
                return 1
        else:
            if ext == '.xlsx':
                if fmtfile == 'JSON':
                    print("Incorrect format file specified for object load")
                    print("File Format detected:", ext)
                    print("File Format given:", ext)
                    self.log(3,
                             "Incorrect format file specified for object load",
                             "File Format detected:", ext,
                             "File Format given:", ext)
                    return 1
                elif fmtfile is None:
                    fmtfile = 'XLSX'
        if fmtobj is None:
            fmtobj = self.fmtobj

        dirname = str(os.path.dirname(dumpfile))
        basename = str(os.path.basename(dumpfile))
        if len(dirname) == 0 and basename == dumpfile:
            dirname = "."
        fileslist = list([])
        my_file = Path(dumpfile)
        if my_file.exists():
            fileslist += [dumpfile]
        else:
            basename = re.sub(ext, "", basename)
            for files in os.listdir(dirname):
                filename = str(files)
                if filename.find(basename) != -1 and \
                   filename.find(ext) != -1:
                    fid = re.sub(".*_", "", re.sub(ext, "", filename))
                    if fid != "None":
                        try:
                            fidint = int(fid)
                        except ValueError:
                            continue
                        if fidint < 0 or fidint > 128:
                            continue
                    fileslist += [str(files)]
        if not any(fileslist):
            print("No dump file found:", dumpfile)
            self.log(2, "No dump file found:", dumpfile)
            return 1
        for i in range(len(fileslist)):
            filename = fileslist[i]
            print("Loading from file :", filename)
            self.log(1, "Loading from file :", filename)
            if fmtfile == 'XLSX':
                wb = openpyxl.load_workbook(filename)
                ws = wb["DUMPHEADER"]
                # A1value = ws['A1'].value
                if fmtfile is None:
                    A2value = ws['A2'].value
                    if str(A2value) in ['XLSX']:
                        fmtfile = str(A2value)
                    else:
                        print("Incorrect Format specified")
                        self.log(1, "Incorrect Format specified")
                        return 1
                if fmtobj is None:
                    A3value = ws['A3'].value
                    if str(A3value) in ['json', 'attributes']:
                        fmtobj = str(A3value)
                    else:
                        fmtobj = self.fmtobj
            else:
                wb = open(filename, 'r')

            # Get the FID
            fid = re.sub(".*_", "", re.sub(ext, "", filename))
            self.loadfids += [fid]

            if fmtfile in 'JSON':
                ret = self.loadfromfid(wb, fid, fmtobj)
                if ret:
                    print("Load failed fmt(", fmtfile, ",", fmtobj, ")")
                    self.log(3, "Load failed fmt(", fmtfile, ",", fmtobj, ")")
                    return ret
            else:
                ws = wb["DUMPHEADER"]
                ret = self.loadfromfidclass(wb, fid, fmtobj, gob)
                if ret:
                    print("Load failed for fmt(", fmtfile, ",", fmtobj, ")")
                    self.log(3, "Load failed fmt(", fmtfile, ",", fmtobj, ")")
                    return ret
                # print(self.loaddict)
                self.log(1, self.loaddict)
        print("Loading Dump configuration Complete.")
        return 0

    def loadfromfid(self, wb, fid, fmtobj):
        """
        Load the switch configuration from a persistent file per fid
        based on JSON format.
        """
        if fmtobj != 'json':
            print("Only JSON Format is supported for the file format")
            return 1
        ws = wb.read()
        wb.close()
        try:
            loadallcell = json.loads(ws)
        except ValueError as err:
            print("Unable to load JSON dump file : ", err)
            self.log(3, "Unable to JSON dump file : ", err)
            return 1
        self.loaded = list(map(lambda x:
                               clsmanager.getInstacefromContainer(
                                   next(iter(x)), x),
                               loadallcell))
        tmpdict = dict()
        for i in range(len(self.clslist)):
            tmpdict.update(dict({self.clslist[i]().getcontainer(): list()}))
        self.loaddict.update(dict({fid: tmpdict}))
        for i in range(len(self.loaded)):
            if self.loaded[i].__class__ not in self.clsfilters:
                self.clsfilters += [self.loaded[i].__class__]
            self.loaddict[fid][self.loaded[i].
                               getcontainer()] += [self.loaded[i]]
        # print (self.loaddict)
        return 0

    def loadmygobjson(self, dumpfile, container):
        """
        Load the switch configuration from a persistent file per fid
        based on JSON format.
        """
        wb = open(dumpfile, "r")
        ws = wb.read()
        wb.close()
        try:
            loadallcell = json.loads(ws)
        except ValueError as err:
            print("Unable to load JSON dump file for container : ", err)
            self.log(3, "Unable to load JSON dump file container : ", err)
            return False, None
        self.loaded = list(map(lambda x:
                               clsmanager.getInstacefromContainer(
                                   next(iter(x)), x),
                               loadallcell))
        for i in range(len(self.loaded)):
            if self.loaded[i].getcontainer() == container:
                return True, self.loaded[i]
        return False, None

    def loadfromfidclass(self, wb, fid, fmtobj, gob):
        """
        Load the switch configuration from a persistent .xlsx per fid
        based on format.
        """
        tmpdict = dict()
        for i in range(len(self.clslist)):
            tmpdict.update(dict({self.clslist[i]().getcontainer(): list()}))
            self.loaddict.update(dict({fid: tmpdict}))

        print("Loading objects fmt type:", fmtobj)
        for i in range(len(self.clslist)):
            container = self.clslist[i]().getcontainer()
            if gob is not None and container != gob:
                continue
            sheetdict = clsmanager.getmultiplesheetdict(container)
            if sheetdict is None or fmtobj == 'json':
                if container not in wb.sheetnames:
                    continue
                ws = wb[container]
            if self.clslist[i] not in self.clsfilters:
                self.clsfilters += [self.clslist[i]]

            if fmtobj == 'json':
                listlen = int(ws.cell(row=1, column=1).value)
                chunk = self.chunkcount
                cols = self.columncount
                loopcount = int((listlen / chunk))
                if chunk > 1 and int(listlen % chunk) > 0:
                    loopcount += 1
                loadcell = list()
                loadallcell = list()
                done = False
                i = 0
                while not done:
                    r = int(2 + (i / cols))
                    c = int(2 + (i % cols))
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None:
                        loadcell = json.loads(cell.value)
                        loadallcell += loadcell
                    else:
                        break
                    i += 1
                # While loop Ends
                self.loaded = list(map(lambda x:
                                       clsmanager.getInstacefromContainer(
                                           next(iter(x)), x),
                                       loadallcell))
                if container == blade().getcontainer():
                    self.loadslots = clsmanager.getbladeslots(self.loaded)
                self.loaddict[fid][container] += self.loaded
            else:
                self.loaded = []
                obj = clsmanager.getInstacefromContainer(container)
                # obj.setdbg_mode(1)
                if sheetdict is None:
                    self.loaded = obj.getmywslist(ws, 5)
                else:
                    obj.setdbg_mode(1)
                    for k, v in sheetdict.items():
                        mysheet = container + "." + k
                        if mysheet in wb.sheetnames:
                            ws = wb[mysheet]
                            self.loaded = obj.getmywslist(ws, 5, v)

                if container == blade().getcontainer():
                    self.loadslots = clsmanager.getbladeslots(self.loaded)
                self.loaddict[fid][container] += self.loaded
        return 0

    def getmyobjectfromrow(self, wb, container, osr):
        """
        Load a specific switch configuration object from a persistent file in
        .xlsx based on container name and start row in attributes format.
        """
        obj = None
        print("Load my row object :", container)
        self.log(1, "Load my row object :", container)
        if container not in wb.sheetnames:
            print("No sheet available with the container name")
            self.log(3, "No sheet available with the container name")
            return False, obj
        obj = clsmanager.getInstacefromContainer(container)
        ws = wb[container]
        ret = obj.loadmyobjfromrow(ws, 5, osr)
        return (ret, obj)

    def addtoslotlist(self, op, indict, inlist):
        """
        Add the objects per slot in list based on operation
        """
        ormop = None
        if op in ("PRE_PATCH", "POST_PATCH", "PRE_DELETE"):
            ormop = clsmanager.getConfigOperation(op)
        for i in range(len(inlist)):
            obj = inlist[i]
            ret, slot = clsmanager.getSlotforObject(obj)
            if ret is False and slot != '0':
                print(obj)
                self.log(3, obj)
            clsmanager.sanitize(obj.__class__, obj)
            if ormop is None:
                if op == 'PATCH':
                    indict[op]['PRE_PATCH'][slot] += [obj]
                else:
                    indict[op][slot] += [obj]
            else:
                indict[ormop][op][slot] += [obj]

    def diff(self, cls, fid, old, new):
        """
        Do a diff between old and new lists for a given class objects
        for a given fid and identify the operations to be performed.
        """
        tmpdict = dict()
        tmpdict.update(dict({'DELETE': dict()}))
        tmpdict.update(dict({'POST': dict()}))
        tmpdict.update(dict({'PATCH': dict()}))
        tmpdict.update(dict({'PUT': dict()}))
        diffset = list()
        onlyold = []
        onlynew = []
        post = clsmanager.getpostsupported(cls)
        for i in range(len(self.diffslots)):
            slot = self.diffslots[i]
            tmpdict['DELETE'].update(dict({slot: list()}))
            tmpdict['POST'].update(dict({slot: list()}))
            tmpdict['PATCH'].update(dict({'PRE_PATCH': dict()}))
            tmpdict['PATCH'].update(dict({'POST_PATCH': dict()}))
            tmpdict['PATCH'].update(dict({'PRE_DELETE': dict()}))
            tmpdict['PATCH']['PRE_PATCH'].update(dict({slot: list()}))
            tmpdict['PATCH']['POST_PATCH'].update(dict({slot: list()}))
            tmpdict['PATCH']['PRE_DELETE'].update(dict({slot: list()}))
            tmpdict['PUT'].update(dict({slot: list()}))

        if cls().getcontainer() == 'n-port-map':
            self.addtoslotlist('DELETE', tmpdict, new)
            self.addtoslotlist('PRE_PATCH', tmpdict, old)
            onlyoldsame = []
            onlynewsame = []
        elif cls().getcontainer() == 'port-group':
            self.addtoslotlist('DELETE', tmpdict, new)
            self.addtoslotlist('POST', tmpdict, old)
            newlbmode = 0
            for items in new:
                if items.peek_port_group_name() == 'pg0':
                    onlynewsame = [items]
                    newlbmode = getattr(items,
                                        "peek_port_group_mode_load" +
                                        "_balancing_mode_enabled")()
                    if newlbmode == 1:
                        tmpobj1 = cls()
                        tmpobj1.set_port_group_id(items.peek_port_group_id())
                        getattr(tmpobj1,
                                "set_port_group_mode_load" +
                                "_balancing_mode_enabled")(0)
                        self.addtoslotlist('PRE_DELETE', tmpdict, [tmpobj1])
                        getattr(items,
                                "set_port_group_mode_load" +
                                "_balancing_mode_enabled")(0)
            for items in old:
                if items.peek_port_group_name() == 'pg0':
                    onlyoldsame = [items]

        elif not clsmanager.isSingletonClass(cls):
            if len(old) == 0:
                self.addtoslotlist('DELETE', tmpdict, new)
            else:
                onlynew = [item for item in new if item not in old]
                self.addtoslotlist('DELETE', tmpdict, onlynew)

            if len(new) == 0:
                self.addtoslotlist(post, tmpdict, old)
            else:
                onlyold = [item for item in old if item not in new]
                self.addtoslotlist(post, tmpdict, onlyold)

            onlyoldsame = [item for item in old if item in new]
            onlynewsame = [item for item in new if item in old]
        else:
            onlyoldsame = old
            onlynewsame = new

        for i in range(len(onlyoldsame)):
            for j in range(len(onlynewsame)):
                oldinst = onlyoldsame[i]
                newinst = onlynewsame[j]
                if oldinst == newinst or clsmanager.isSingletonClass(cls):
                    if cls().getcontainer() == 'effective-configuration':
                        if oldinst.peek_default_zone_access() is not None and\
                           newinst.peek_default_zone_access() != \
                           oldinst.peek_default_zone_access():
                            obj = cls()
                            obj.set_default_zone_access(
                                oldinst.peek_default_zone_access())
                            obj.set_cfg_action(1)
                            self.addtoslotlist('PRE_PATCH', tmpdict, [obj])
                        if newinst.peek_cfg_name() is not None:
                            obj = cls()
                            obj.set_cfg_action(2)
                            self.addtoslotlist('PRE_DELETE', tmpdict,
                                               [obj])
                        if oldinst.peek_cfg_name() is not None:
                            obj = cls()
                            obj.set_cfg_name(oldinst.peek_cfg_name())
                            self.addtoslotlist('POST_PATCH', tmpdict, [obj])
                    else:
                        diffset.append((oldinst, newinst))
                        oldinst.diff(newinst, clsmanager.diffignoreattrib(cls))
                        # Add stage ordering with Diff
                        clsmanager.stageordering(cls, oldinst, newinst)
                        # print ("CONFIGCHANGED old:", oldinst.configchanged,
                        #       "new : ", newinst.configchanged,
                        #       oldinst, newinst,
                        #       oldinst.create_html_content(0, None),
                        #       newinst.create_html_content(5, None))
                        if (oldinst.configchanged & 4) == 4:
                            self.addtoslotlist('POST', tmpdict, [oldinst])
                        if (oldinst.configchanged & 2) == 2:
                            if cls().getcontainer() == 'fibrechannel-switch':
                                self.addtoslotlist('PRE_DELETE',
                                                   tmpdict, [oldinst])
                            else:
                                self.addtoslotlist('PATCH', tmpdict, [oldinst])
                        if (newinst.configchanged & 8) == 8:
                            self.addtoslotlist('DELETE', tmpdict, [newinst])
        if fid not in self.diffdict.keys():
            self.diffdict.update(dict({fid: dict()}))
        if fid not in self.diffdetails.keys():
            self.diffdetails.update(dict({fid: dict()}))
        # print(fid, cls, tmpdict)
        self.diffdetails[fid].update(dict({cls().getcontainer():
                                           dict({'oldonlylist': onlyold,
                                                 'newonlylist': onlynew,
                                                 'oldlist': old,
                                                 'newlist': new,
                                                 'samelist': diffset})}))
        self.diffdict[fid].update(dict({cls().getcontainer(): tmpdict}))

    def showdiff(self, session=None):
        """
        Display the diff calculation after the diff operation is completed.
        """
        # utils.response_print(self.diffdict)
        ADD = '+++'
        DEL = '---'
        MOD = '***'
        alldiff = list()
        for i in range(len(self.clslist)):
            cls = self.clslist[i]
            for j in range(len(self.difffids)):
                fid = self.difffids[j]
                allop = clsmanager.getConfigStages()
                for k in range(len(allop)):
                    op = allop[k]
                    for l in range(len(self.diffslots)):
                        slot = self.diffslots[l]
                        entry = self.get(fid, cls, op, slot, session)
                        if len(entry) == 0:
                            continue
                        tmpdict = dict(dict({str(cls):
                                            dict({fid:
                                                  dict({slot:
                                                        list()})})}))
                        trydiff = dict()
                        tmpdict[str(cls)][fid][slot] += entry
                        ormop = clsmanager.getConfigOperation(op)
                        if ormop == 'POST':
                            trydiff[ADD] = entry
                        if ormop == 'DELETE':
                            trydiff[DEL] = entry
                        if ormop in ('PATCH', 'PUT'):
                            for items in entry:
                                if items.configchanged == 0:
                                    trydiff[ADD] = [items]
                                else:
                                    trydiff[MOD] = [items]

                        alldiff += [trydiff]
        stralldiff = utils.strjson(alldiff)
        linuxdiff = ""
        exp_pattern = ""
        marker = " "
        revpattern = ""
        popthenext = 0
        for lines in stralldiff.splitlines():
            if popthenext == 1:
                popthenext = 0
                continue
            if len(exp_pattern) == 0:
                marker = "   "
            if re.search(r"\+\+\+", lines):
                marker = '> +'
                if re.search(r"\"\+\+\+\": \[", lines):
                    exp_pattern = re.sub(r"\"\+\+\+\": \[", "]", str(lines))
            if re.search(r"^.*\".*\": \{", lines):
                attribute = re.sub(r"[\ ]*\"", "",
                                   re.sub(r"\": \{", "", str(lines)))
            # list element marked for diff
            if re.search(r"\"\[\+]", lines):
                marker = '> +'
                exp_pattern = re.sub(r"    \".*,", "}", str(lines))
                revpattern = re.sub(r"    \".*,", r"\{", str(lines))
                oldline = linuxdiff.splitlines()
                i = len(oldline) - 1
                while i > 0:
                    oldline[i] = re.sub(r"^.", marker, oldline[i])
                    if re.search(revpattern, oldline[i]):
                        # match from starting to end instead of
                        # caret use length
                        if len(oldline[i]) == len(revpattern):
                            break
                    i = i - 1
                linuxdiff = "\n".join(oldline) + "\n"
            if re.search(r"---", lines):
                marker = '< -'
                if re.search(r"\"---\": \[", lines):
                    exp_pattern = re.sub(r"\"---\": \[", "]", str(lines))
            # Object list element marked for diff
            if re.search(r"\"\[-\]", lines):
                marker = '< -'
                exp_pattern = re.sub(r"    \".*,", "}", str(lines))
                revpattern = re.sub(r"    \".*,", r"\{", str(lines))
                oldline = linuxdiff.splitlines()
                i = len(oldline) - 1
                while i > 0:
                    oldline[i] = re.sub("^.", marker, oldline[i])
                    if re.search(revpattern, oldline[i]):
                        # match from starting to end instead of
                        # caret use length
                        if len(oldline[i]) == len(revpattern):
                            break
                    i = i - 1
                linuxdiff = "\n".join(oldline) + "\n"

            if re.search("Original", lines):
                marker = '> +'
                lines = re.sub(r"\ \ \ \ \"Original",
                               r"" + str("\"") + attribute, lines)
                linuxdiff += (marker + lines + "\n")
                popthenext = 1
            elif re.search("Drifted", lines):
                oldline = linuxdiff.splitlines()
                oldline.pop()
                linuxdiff = "\n".join(oldline) + "\n"
                marker = '< -'
                lines = re.sub(",$", "",
                               re.sub(r"\ \ \ \ \"Drifted",
                                      r"" + str("\"") + attribute, lines))
                linuxdiff += (marker + lines + "\n")
                linuxdiff += "---\n"
            else:
                linuxdiff += (marker + lines + "\n")
            # match from starting to end
            if len(exp_pattern) and re.search("^" + exp_pattern, lines):
                exp_pattern = ""
        linuxdiff = re.sub(r"\"...\": \[", "[", linuxdiff)
        linuxdiff = re.sub(r"\[[+|-]]", "", linuxdiff)
        # linuxdiff = re.sub("\"\+\+\+\": \"", "\"Original\": \"", linuxdiff)
        # linuxdiff = re.sub("\"---\": \"", "\"Drifted\": \"", linuxdiff)
        self.log(1, "Stage diff \n", utils.strjson(self.diffdict))
        self.log(1, "List diff diff \n", utils.strjson(alldiff))
        self.log(1, "Linux diff \n", linuxdiff)
        # print(utils.strjson(alldiff))
        utils.initscreen()
        for lines in linuxdiff.splitlines():
            if re.search(r"^< -", lines):
                utils.print_test_red(lines)
            elif re.search(r"^> \+", lines):
                utils.print_test_green(lines)
            else:
                utils.print_test_blue(lines)
        print(linuxdiff)

    def handlediff(self, file, session):
        """
        Start the diff handling for all the switch configuration.
        """
        print("Handle Diff Start.")
        self.log(1, "Handle Diff Start.")
        ret = self.loadfromfile(file)
        if ret:
            self.log(3, "Load from file Failed")
            print("Load from file Failed")
            return ret
        self.loadcurrent(session)
        self.diffslots = list(set(self.currentslots + self.loadslots))
        self.difffids = list(set(self.loadfids + self.currentfids))
        print("Init Class Ordering.")
        self.log(1, "Init Class Ordering.")
        clsmanager.clsorderinginit(self.diffslots)
        print("Calculating Diff Start")
        self.log(1, "Calculating Diff Start")
        for j in range(len(self.difffids)):
            fid = self.difffids[j]
            sfid = fid
            dfid = fid
            fidkeys = self.currentdict.keys()
            if fid not in fidkeys:
                if fid == "None":
                    if '128' in fidkeys:
                        dfid = '128'

            if fid not in self.loadfids:
                if fid == "None":
                    if '128' in self.loadfids:
                        sfid = '128'
            self.diffdict.update(dict({fid: dict()}))
            self.diffdetails.update(dict({fid: dict()}))
            for i in range(len(self.clslist)):
                tgtcls = self.clslist[i]
                if tgtcls not in self.clsfilters:
                    continue
                clskey = tgtcls().getcontainer()
                if fid != '128':
                    if fid != 'None':
                        if not clsmanager.checkvfsupported(tgtcls):
                            continue
                if clsmanager.isConfigClass(tgtcls, session):
                    if sfid in self.loaddict.keys() and \
                       clskey in self.loaddict[sfid].keys():
                        old = self.loaddict[sfid][clskey]
                    else:
                        old = []
                    if dfid in self.currentdict.keys() and \
                       clskey in self.currentdict[dfid].keys():
                        new = self.currentdict[dfid][clskey]
                    else:
                        new = []
                    self.diff(self.clslist[i], fid, old, new)
        print("Calculating Diff Complete.")
        self.log(1, "Calculating Diff Complete.")
        self.showdiff()
        return 0

    def applydiff(self, file, session):
        """
        Apply the diff handling for all the switch configuration.
        """
        ret = self.handlediff(file, session)
        if ret:
            print("Load or Diff of configs Failed")
            self.log(3, "Load or Diff of configs Failed")
            return ret
        print("Init Rule Manager.")
        self.log(1, "Init Rule Manager.")
        self.rmmanager = rmmanager(rmmanager.defaultconfigrules(), self)
        # self.rmmanager.displayclsorm()
        print("Apply Rule Manager.")
        self.log(1, "Apply Rule Manager.")
        self.rmmanager.allconfigrules(session, self.difffids, self.diffslots)
        # print (self.diffdict)
        return 0

    def applygoldenobject(self, session, dumpfile, container, osr):
        """
        Apply a specific golden configuration object from a persistent fil
        to all the objects of the same class on the switch.
        """
        my_file = Path(dumpfile)
        if not my_file.exists():
            # print("No file found with the name", dumpfile)
            self.log(1, "No file found with the name", dumpfile)
            return False
        # pylint: disable=W1401
        ext = re.sub(".*\.", ".", str(dumpfile))
        if ext == '.json':
            ret, gob = self.loadmygobjson(dumpfile, container)
            print(gob, container)
        else:
            wb = openpyxl.load_workbook(dumpfile)
            ret, gob = self.getmyobjectfromrow(wb, container, osr)

        if ret is False:
            print("Unable to load the object")
            self.log(3, "Unable to load the object",
                     dumpfile, container, osr)
            return False
        # print(gob)
        self.log(1, dumpfile, container, osr, gob)
        clsmanager.sleep(10, session)
        self.loadcurrent(session, container)
        clsmanager.clsorderinginit(self.currentslots)
        self.diffslots = self.currentslots
        self.difffids = self.currentfids
        self.rmmanager = rmmanager(rmmanager.defaultconfigrules(), self)
        for j in range(len(self.currentfids)):
            fid = self.currentfids[j]
            self.diffdict.update(dict({fid: dict()}))
            self.loaddict.update(dict({fid: dict()}))
            self.diffdetails.update(dict({fid: dict()}))
            for i in range(len(self.clslist)):
                clskey = self.clslist[i]().getcontainer()
                if clskey != container:
                    continue
                if clsmanager.isConfigClass(self.clslist[i], session):
                    new = self.currentdict[fid][clskey]
                    # Form the old list
                    oldlist = []
                    for m in range(len(new)):
                        loadconfig = json.loads(gob.__repr__())
                        newobj = self.clslist[i](loadconfig)
                        for k in range(len(newobj.keyslist)):
                            for l in range(len(new[m].keyslist)):
                                if newobj.keyslist[k].uname ==\
                                   new[m].keyslist[l].uname:
                                    newobj.keyslist[k].setvalue(
                                        new[m].keyslist[l].value)
                        if container in 'fibrechannel':
                            newobj.set_user_friendly_name(
                                new[m].peek_user_friendly_name())
                        oldlist.append(newobj)
                    self.loaddict[fid].update(dict({container: oldlist}))
                    # End of old list creation
                    self.diff(self.clslist[i], fid, oldlist, new)
        self.showdiff(session)
        print("Apply Rule Manager.")
        self.log(1, "Apply Rule Manager.")
        self.rmmanager.allconfigrules(session,
                                      self.currentfids,
                                      self.currentslots)
        return True

    def get(self, fid, cls, op, slot, session):
        """
        Get Handler for diff configuration to work with Rule Manager.
        """
        if fid == -1:
            fid = "None"
        ormop = clsmanager.getConfigOperation(op)
        if clsmanager.isConfigClass(cls, session):
            if fid in self.diffdict.keys():
                container = cls().getcontainer()
                if container in self.diffdict[fid].keys():
                    if ormop in self.diffdict[fid][container].keys():
                        if ormop == 'PATCH':
                            if op in \
                               self.diffdict[fid][container][ormop].keys():
                                if slot in\
                                   self.diffdict[fid][container][ormop][op].\
                                   keys():
                                    tmp = self.diffdict[fid][container]
                                    return tmp[ormop][op][slot]
                        else:
                            if slot in\
                               self.diffdict[fid][container][ormop].keys():
                                tmp = self.diffdict[fid][container]
                                return tmp[ormop][slot]
        # print("No config for :FID", fid, " XLSX:", cls,
        #      " OP: ", ormop, " Slot:", slot)
        return list([])

    def handleCB(self, cls, op, slot, session):
        """
        Callback Handler to integrate the session reinit handling and sleep
        handling to work with Rule Manager.
        """
        CB = clsmanager.getClsStageCB(cls, op, slot)
        if CB is not None:
            # print("CB is set for=>", cls, op, slot)
            self.log(1, "CB is set for=>", cls, op, slot)
            clsmanager.cbhandler(session, CB)
